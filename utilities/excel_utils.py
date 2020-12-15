import openpyxl


def get_row_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def get_column_count(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def read_data(file, sheetName, row_number, column_number):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=row_number, column=column_number).value


def write_data(file, sheetName, row_number, column_number, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(file)
